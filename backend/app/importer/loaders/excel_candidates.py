"""Excel candidate loader — ports the legacy `load_excel_data()` from
`election_dashboard.py:233` into the new schema.

Reads the "Election Overview" and "Chairmanship Candidates" sheets from
the FCT 2026 workbook. Generalized to any state by passing `--state` to the
CLI; the same loader can ingest similar workbooks for other states.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from sqlalchemy import select

from app.db import session_scope
from app.importer.normalizers import resolve_party
from app.importer.schemas import ImportSummary
from app.models import Candidate, Election, IngestionSource, Lga, State
from app.scraper.phases import ensure_election

log = logging.getLogger(__name__)


def load_excel(*, filepath: Path, cycle: int, state_code: str) -> ImportSummary:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    rows_in = 0
    rows_imported = 0
    rows_skipped = 0
    errors: list[str] = []
    unmapped: set[str] = set()
    elections_touched: set[int] = set()

    with session_scope() as session:
        state = session.scalar(select(State).where(State.code == state_code))
        if state is None:
            return ImportSummary(
                rows_in=0,
                rows_imported=0,
                rows_skipped=0,
                elections_touched=0,
                unmapped_parties=[],
                errors=[f"unknown state_code={state_code}"],
            )

        source = session.scalar(
            select(IngestionSource).where(IngestionSource.name == f"excel_candidates_{cycle}")
        )
        if source is None:
            source = IngestionSource(
                name=f"excel_candidates_{cycle}",
                url=str(filepath),
                license="public",
                notes="Ported from legacy load_excel_data()",
            )
            session.add(source)
            session.flush()

        # Sheet 1: Election Overview — area councils / LGAs
        if "Election Overview" in wb.sheetnames:
            ws = wb["Election Overview"]
            for row in ws.iter_rows(min_row=5, max_row=200, values_only=True):
                if not row or not row[0]:
                    break
                lga_name = str(row[0]).strip()
                existing = session.scalar(
                    select(Lga).where(Lga.state_id == state.state_id, Lga.name == lga_name)
                )
                if existing is None:
                    session.add(
                        Lga(state_id=state.state_id, name=lga_name, lga_kind="area_council")
                    )

        # Sheet 2: Chairmanship Candidates
        if "Chairmanship Candidates" in wb.sheetnames:
            ws = wb["Chairmanship Candidates"]
            for row in ws.iter_rows(min_row=4, max_row=200, values_only=True):
                if not row or not row[1]:
                    continue
                rows_in += 1
                _, area_council, candidate_name, _party_full, party_abbrev, _status, _gender, _notes = (
                    row + (None,) * 8
                )[:8]
                if not (candidate_name and party_abbrev):
                    rows_skipped += 1
                    continue

                party = resolve_party(session, code=str(party_abbrev), cycle=cycle, autocreate=True)
                if party is None:
                    unmapped.add(str(party_abbrev))
                    rows_skipped += 1
                    continue

                election = ensure_election(
                    session,
                    cycle=cycle,
                    election_type="lg_chairman",
                    state_id=state.state_id,
                    irev_election_id=None,
                    election_date=None,
                    status="historical",
                )
                elections_touched.add(election.election_id)

                existing_cand = session.scalar(
                    select(Candidate).where(
                        Candidate.election_id == election.election_id,
                        Candidate.party_id == party.party_id,
                    )
                )
                if existing_cand is None:
                    session.add(
                        Candidate(
                            election_id=election.election_id,
                            party_id=party.party_id,
                            full_name=str(candidate_name).strip(),
                            is_incumbent=False,
                        )
                    )
                    rows_imported += 1

    return ImportSummary(
        rows_in=rows_in,
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
        elections_touched=len(elections_touched),
        unmapped_parties=sorted(unmapped),
        errors=errors[:50],
    )
